import openpyxl

# Function to read the UUIDs from the xlsx file
def read_uuids_from_xlsx(file_name):
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active

    uuids = [row[0] for row in sheet.iter_rows(min_row=2, values_only=True)]
    return set(uuids)

# Function to read the UUIDs from the txt file
def read_uuids_from_txt(file_name):
    with open(file_name, 'r') as file:
        lines = file.readlines()
    
    # Extract UUID from each line
    uuids = [line.rstrip('/\n').split('/')[-1] for line in lines]
    return set(uuids)

# Function to write the missing UUIDs to a new xlsx file
def write_uuids(file_name, uuids):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    for uuid in uuids:
        sheet.append([uuid])

    workbook.save(file_name)

# Main script

# Read the UUIDs from the xlsx file
xlsx_uuids = read_uuids_from_xlsx('active-buckets-27.06.2023.xlsx')

# Read the UUIDs from the txt file
bucket_uuids = read_uuids_from_txt('bucket_uuids.txt')

# Get the UUIDs that are in the bucket but not in the xlsx file
missing_uuids = bucket_uuids - xlsx_uuids

# Write the missing UUIDs to a new xlsx file
write_uuids('missing-uuids.xlsx', missing_uuids)

print(f'Process completed. {len(missing_uuids)} missing UUIDs written to missing-uuids.xlsx')
