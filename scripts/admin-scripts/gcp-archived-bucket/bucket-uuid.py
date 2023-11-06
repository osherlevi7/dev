from google.cloud import storage
import openpyxl

# Function to list all the UUIDs in the GCP bucket
def list_uuids(bucket_name, path):
    storage_client = storage.Client()
    bucket = storage_client.bucket(bucket_name)
    blobs = bucket.list_blobs(prefix=path)

    uuids = []
    for blob in blobs:
        uuid = blob.name.replace(path, "").split('/')[0]
        print(f'Blob name: {blob.name}, Extracted UUID: {uuid}')  # Debug print
        if uuid:
            uuids.append(uuid)

    return set(uuids)  # return set for faster lookup


# Function to read the UUIDs from the xlsx file
def read_uuids_from_xlsx(file_name):
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active

    uuids = []
    for row in sheet.iter_rows(values_only=True):
        uuids.append(row[0])  # Assuming UUIDs are in the first column

    return set(uuids)  # return set for faster lookup

# Function to check and write the missing UUIDs to a new xlsx file
def check_and_write_uuids(file_name, uuids_in_bucket, uuids_in_file):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    for uuid in uuids_in_bucket:
        if uuid not in uuids_in_file:
            sheet.append([uuid])

    workbook.save(file_name)

# Main script

# Specify your bucket name
bucket_name = "scans" 
# Specify the path where the UUIDs are stored in the bucket
path = "/webroot/union/"


# List the UUIDs in the bucket
uuids_in_bucket = list_uuids(bucket_name, path)

# Specify the name of the existing xlsx file
file_name = "active-buckets-27.06.2023.xlsx"

# Read the UUIDs from the xlsx file
uuids_in_file = read_uuids_from_xlsx(file_name)

# Specify the name of the new xlsx file to be created
new_file_name = "missing-uuids.xlsx"

# Check and write the missing UUIDs to a new xlsx file
check_and_write_uuids(new_file_name, uuids_in_bucket, uuids_in_file)

print("Process completed. Missing UUIDs written to: ", new_file_name)
